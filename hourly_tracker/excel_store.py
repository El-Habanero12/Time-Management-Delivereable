from __future__ import annotations

import os
import tempfile
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Union

import msvcrt
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hourly_tracker.first_run import ensure_user_files_exist
ENTRIES_SHEET = "Entries"
LOOKUP_SHEET = "Lookup"
TASKS_SHEET = "Tasks"
TASK_EVENTS_SHEET = "Task_Events"
REFLECTIONS_INDEX_SHEET = "Reflections_Index"

ENTRIES_COLUMNS = [
    "id",
    "timestamp",
    "activity",
    "notes",
    "category",
    "energy",
    "focus",
    "prompt_type",
    "start_time",
    "end_time",
    "created_at",
]

LOOKUP_COLUMNS = [
    "category",
    "keywords",
    "regex",
    "updated_at",
]

TASKS_COLUMNS = [
    "id",
    "title",
    "status",
    "created_at",
    "completed_at",
    "last_worked_at",
    "total_minutes",
    "notes",
]

TASK_EVENTS_COLUMNS = [
    "id",
    "task_id",
    "timestamp",
    "action",
    "minutes",
    "effort",
    "could_be_faster",
    "notes",
]

DEFAULT_CATEGORIES = [
    "Work",
    "Study",
    "Admin",
    "Break",
    "Exercise",
    "Chores",
    "Social",
    "Other",
]


def _as_path(p: Union[str, Path]) -> Path:
    return p if isinstance(p, Path) else Path(p)


def _lock_path(path: Union[str, Path], override: Optional[Union[str, Path]]) -> Path:
    target = _as_path(path)
    lock = _as_path(override) if override is not None else target.with_suffix(target.suffix + ".lock")
    target.parent.mkdir(parents=True, exist_ok=True)
    lock.parent.mkdir(parents=True, exist_ok=True)
    return lock


def _norm_header(value: object) -> str:
    """Normalize header names so 'Payment Method' and 'payment_method' match."""
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


@dataclass
class Entry:
    timestamp: datetime
    activity: str
    notes: str = ""
    category: str = "Other"
    energy: int = 3
    focus: int = 3
    prompt_type: str = "regular"
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    entry_id: Optional[str] = None

    def to_row_dict(self) -> Dict[str, object]:
        now = datetime.now()
        return {
            "id": self.entry_id or f"{int(now.timestamp() * 1000)}",
            "timestamp": self.timestamp.isoformat(timespec="seconds"),
            "activity": self.activity,
            "notes": self.notes,
            "category": self.category,
            "energy": int(self.energy),
            "focus": int(self.focus),
            "prompt_type": self.prompt_type,
            "start_time": self.start_time.isoformat(timespec="seconds") if self.start_time else "",
            "end_time": self.end_time.isoformat(timespec="seconds") if self.end_time else "",
            "created_at": now.isoformat(timespec="seconds"),
        }


@contextmanager
def file_lock(lock_path: Path, timeout_seconds: float = 5.0, poll_seconds: float = 0.1):
    """Simple Windows file lock using msvcrt on a .lock file."""
    lock_path = _as_path(lock_path)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    start = time.time()
    with open(lock_path, "a+b") as lock_file:
        while True:
            try:
                lock_file.seek(0)
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
                break
            except OSError:
                if (time.time() - start) >= timeout_seconds:
                    raise TimeoutError(f"Timed out waiting for lock: {lock_path}")
                time.sleep(poll_seconds)
        try:
            yield
        finally:
            lock_file.seek(0)
            try:
                msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
            except OSError:
                # If unlocking fails, there isn't much we can safely do.
                pass


def _atomic_save(wb: Workbook, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=target_path.stem + "_", suffix=".tmp", dir=str(target_path.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, target_path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def atomic_save_workbook(wb: Workbook, target_path: Path) -> None:
    _atomic_save(wb, target_path)


def _ensure_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _get_header_map(ws: Worksheet, expected_columns: List[str]) -> Dict[str, int]:
    header_row = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    header_row = [str(h) if h is not None else "" for h in header_row]

    if not any(header_row):
        ws.append(expected_columns)
        return {col: idx + 1 for idx, col in enumerate(expected_columns)}

    header_set = set(header_row)
    missing = [col for col in expected_columns if col not in header_set]
    if missing:
        # Extend header row with missing stable columns.
        for col in missing:
            header_row.append(col)
        for idx, col in enumerate(header_row, start=1):
            ws.cell(row=1, column=idx, value=col)

    return {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}


def _find_header_row(ws: Worksheet, expected_columns: List[str], max_scan: int = 5) -> Optional[int]:
    for row in range(1, min(max_scan, ws.max_row) + 1):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row]]
        if not values:
            continue
        matches = sum(1 for col in expected_columns if col in values)
        if matches >= max(2, len(expected_columns) // 2):
            return row
    return None


def _ensure_lookup_categories(ws: Worksheet, categories: Iterable[str]) -> None:
    header_map = _get_header_map(ws, LOOKUP_COLUMNS)
    category_col = header_map["category"]

    existing = set()
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=category_col).value
        if value:
            existing.add(str(value))

    now = datetime.now().isoformat(timespec="seconds")
    for cat in categories:
        if cat not in existing:
            ws.append([cat, "", "", now])


def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    # Rename the default sheet to Entries for consistency.
    default_ws = wb.active
    default_ws.title = ENTRIES_SHEET
    return wb


def load_or_create_workbook(path: Path) -> Workbook:
    return _load_or_create(path)


def ensure_workbook(path: Path, categories: Optional[Iterable[str]] = None, lock_path: Optional[Path] = None) -> None:
    path = _as_path(path)
    categories = list(categories or DEFAULT_CATEGORIES)
    lock = _lock_path(path, lock_path)

    with file_lock(lock):
        wb = _load_or_create(path)
        entries_ws = _ensure_sheet(wb, ENTRIES_SHEET)
        lookup_ws = _ensure_sheet(wb, LOOKUP_SHEET)
        tasks_ws = _ensure_sheet(wb, TASKS_SHEET)
        task_events_ws = _ensure_sheet(wb, TASK_EVENTS_SHEET)
        _get_header_map(entries_ws, ENTRIES_COLUMNS)
        _ensure_lookup_categories(lookup_ws, categories)
        _get_header_map(tasks_ws, TASKS_COLUMNS)
        _get_header_map(task_events_ws, TASK_EVENTS_COLUMNS)
        _atomic_save(wb, path)


def append_entry(path: Path, entry: Entry, categories: Optional[Iterable[str]] = None, lock_path: Optional[Path] = None) -> None:
    path = _as_path(path)
    categories = list(categories or DEFAULT_CATEGORIES)
    lock = _lock_path(path, lock_path)
    row_dict = entry.to_row_dict()

    with file_lock(lock):
        wb = _load_or_create(path)
        entries_ws = _ensure_sheet(wb, ENTRIES_SHEET)
        lookup_ws = _ensure_sheet(wb, LOOKUP_SHEET)

        header_map = _get_header_map(entries_ws, ENTRIES_COLUMNS)
        _ensure_lookup_categories(lookup_ws, categories)

        row_values = [""] * len(header_map)
        for col_name, col_idx in header_map.items():
            if col_name in row_dict:
                row_values[col_idx - 1] = row_dict[col_name]

        entries_ws.append(row_values)
        _atomic_save(wb, path)


def read_entries(path: Path, lock_path: Optional[Path] = None) -> List[Dict[str, object]]:
    path = _as_path(path)
    if not path.exists():
        return []

    lock = _lock_path(path, lock_path)
    with file_lock(lock, timeout_seconds=5.0):
        wb = load_workbook(path, data_only=True)
        if ENTRIES_SHEET not in wb.sheetnames:
            return []
        ws = wb[ENTRIES_SHEET]
        if ws.max_row < 1:
            return []

        def find_header_row() -> tuple[Optional[int], List[str]]:
            # Prefer row 1; fallback scan first 20 rows for timestamp + activity.
            candidates = [1] + list(range(1, min(20, ws.max_row) + 1))
            seen = set()
            for idx in candidates:
                if idx in seen:
                    continue
                seen.add(idx)
                headers = [_norm_header(c.value) for c in ws[idx]]
                if any(headers):
                    if "timestamp" in headers and "activity" in headers:
                        return idx, headers
            return None, []

        header_row_idx, headers = find_header_row()
        if not header_row_idx or not headers:
            # Explicit log via caller; return empty to halt analytics safely.
            return []

        results: List[Dict[str, object]] = []
        read_entries.last_headers = headers  # type: ignore[attr-defined]
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v is None for v in row):
                continue
            row_dict: Dict[str, object] = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                row_dict[h] = v
            if row_dict:
                results.append(row_dict)
        return results


def read_categories(path: Path, lock_path: Optional[Path] = None) -> List[str]:
    path = _as_path(path)
    if not path.exists():
        return list(DEFAULT_CATEGORIES)

    lock = _lock_path(path, lock_path)
    with file_lock(lock, timeout_seconds=5.0):
        wb = load_workbook(path, data_only=True)
        if LOOKUP_SHEET not in wb.sheetnames:
            return list(DEFAULT_CATEGORIES)
        ws = wb[LOOKUP_SHEET]
        header_map = _get_header_map(ws, LOOKUP_COLUMNS)
        category_col = header_map["category"]
        cats: List[str] = []
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=category_col).value
            if value:
                cats.append(str(value))
        return cats or list(DEFAULT_CATEGORIES)


def read_tasks(path: Path, status_filter: Optional[str] = "open", lock_path: Optional[Path] = None) -> List[Dict[str, object]]:
    path = _as_path(path)
    if not path.exists():
        return []

    lock = _lock_path(path, lock_path)
    with file_lock(lock, timeout_seconds=5.0):
        wb = load_workbook(path, data_only=True)
        if TASKS_SHEET not in wb.sheetnames:
            return []
        ws = wb[TASKS_SHEET]
        if ws.max_row < 2:
            return []

        header_row = _find_header_row(ws, TASKS_COLUMNS) or 1
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row]]
        results: List[Dict[str, object]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            status_value = str(row_dict.get("status") or "").strip().lower()
            if status_filter and status_value != status_filter:
                continue
            results.append(row_dict)
        return results


def read_task_events(path: Path, lock_path: Optional[Path] = None) -> List[Dict[str, object]]:
    path = _as_path(path)
    if not path.exists():
        return []

    lock = _lock_path(path, lock_path)
    with file_lock(lock, timeout_seconds=5.0):
        wb = load_workbook(path, data_only=True)
        if TASK_EVENTS_SHEET not in wb.sheetnames:
            return []
        ws = wb[TASK_EVENTS_SHEET]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        headers = [_norm_header(h) for h in header_row]
        header_set = set(headers)
        results: List[Dict[str, object]] = []
        for row in rows:
            if all(v is None for v in row):
                continue  # skip empty rows
            # Skip if this row looks like a repeated header row (any cell matches a header name).
            if any((_norm_header(v) in header_set) for v in row if v is not None):
                continue
            row_dict: Dict[str, object] = {}
            for h, v in zip(headers, row):
                if not h:
                    continue  # skip blank header columns so headers never appear as data
                row_dict[h] = v
            if row_dict:
                results.append(row_dict)
        return results


def add_tasks(path: Path, titles: Iterable[str], lock_path: Optional[Path] = None) -> List[str]:
    path = _as_path(path)
    lock = _lock_path(path, lock_path)
    new_ids: List[str] = []

    with file_lock(lock):
        wb = _load_or_create(path)
        tasks_ws = _ensure_sheet(wb, TASKS_SHEET)
        header_map = _get_header_map(tasks_ws, TASKS_COLUMNS)

        now = datetime.now().isoformat(timespec="seconds")
        for title in titles:
            clean = title.strip()
            if not clean:
                continue
            task_id = f"task_{int(datetime.now().timestamp() * 1000)}_{len(new_ids)}"
            row_values = [""] * len(header_map)
            row_values[header_map["id"] - 1] = task_id
            row_values[header_map["title"] - 1] = clean
            row_values[header_map["status"] - 1] = "open"
            row_values[header_map["created_at"] - 1] = now
            row_values[header_map["total_minutes"] - 1] = 0
            tasks_ws.append(row_values)
            new_ids.append(task_id)

        _atomic_save(wb, path)

    return new_ids


def _find_task_row(ws: Worksheet, task_id: str, id_col: int) -> Optional[int]:
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=id_col).value
        if str(value) == task_id:
            return row
    return None


def update_task_fields(path: Path, task_id: str, title: Optional[str] = None, notes: Optional[str] = None, lock_path: Optional[Path] = None) -> None:
    path = _as_path(path)
    lock = _lock_path(path, lock_path)
    with file_lock(lock):
        wb = _load_or_create(path)
        tasks_ws = _ensure_sheet(wb, TASKS_SHEET)
        tasks_header = _get_header_map(tasks_ws, TASKS_COLUMNS)

        id_col = tasks_header["id"]
        row_idx = _find_task_row(tasks_ws, task_id, id_col)
        if row_idx:
            if title is not None:
                tasks_ws.cell(row=row_idx, column=tasks_header["title"], value=title)
            if notes is not None:
                tasks_ws.cell(row=row_idx, column=tasks_header["notes"], value=notes)
        _atomic_save(wb, path)


def log_task_event(
    path: Path,
    task_id: str,
    action: str,
    minutes: int,
    effort: int,
    could_be_faster: bool,
    notes: str = "",
    lock_path: Optional[Path] = None,
) -> None:
    return log_task_event_with_lock(path, task_id, action, minutes, effort, could_be_faster, notes, lock_path)


def log_task_event_with_lock(
    path: Path,
    task_id: str,
    action: str,
    minutes: int,
    effort: int,
    could_be_faster: bool,
    notes: str = "",
    lock_path: Optional[Path] = None,
) -> None:
    path = _as_path(path)
    lock = _lock_path(path, lock_path)
    with file_lock(lock):
        wb = _load_or_create(path)
        tasks_ws = _ensure_sheet(wb, TASKS_SHEET)
        events_ws = _ensure_sheet(wb, TASK_EVENTS_SHEET)

        tasks_header = _get_header_map(tasks_ws, TASKS_COLUMNS)
        events_header = _get_header_map(events_ws, TASK_EVENTS_COLUMNS)

        now = datetime.now()
        event_id = f"event_{int(now.timestamp() * 1000)}"

        row_values = [""] * len(events_header)
        row_values[events_header["id"] - 1] = event_id
        row_values[events_header["task_id"] - 1] = task_id
        row_values[events_header["timestamp"] - 1] = now.isoformat(timespec="seconds")
        row_values[events_header["action"] - 1] = action
        row_values[events_header["minutes"] - 1] = int(minutes)
        row_values[events_header["effort"] - 1] = int(effort)
        row_values[events_header["could_be_faster"] - 1] = bool(could_be_faster)
        row_values[events_header["notes"] - 1] = notes
        events_ws.append(row_values)

        id_col = tasks_header["id"]
        row_idx = _find_task_row(tasks_ws, task_id, id_col)
        if row_idx:
            if action == "worked":
                last_worked_col = tasks_header["last_worked_at"]
                total_minutes_col = tasks_header["total_minutes"]
                tasks_ws.cell(row=row_idx, column=last_worked_col, value=now.isoformat(timespec="seconds"))
                existing = tasks_ws.cell(row=row_idx, column=total_minutes_col).value or 0
                tasks_ws.cell(row=row_idx, column=total_minutes_col, value=int(existing) + int(minutes))
            elif action == "completed":
                status_col = tasks_header["status"]
                completed_col = tasks_header["completed_at"]
                tasks_ws.cell(row=row_idx, column=status_col, value="done")
                tasks_ws.cell(row=row_idx, column=completed_col, value=now.isoformat(timespec="seconds"))

        _atomic_save(wb, path)


# -------------------- Expenses workbook support --------------------

def _find_expense_headers(ws: Worksheet, max_scan_rows: int = 20) -> tuple[Optional[int], Dict[str, int]]:
    header_row_idx: Optional[int] = None
    header_map: Dict[str, int] = {}
    limit = min(max_scan_rows, max(1, ws.max_row))
    for row_idx in range(1, limit + 1):
        cells = list(ws[row_idx])
        lowered = [str(c.value).strip().lower() for c in cells if c.value is not None]
        if not lowered:
            continue
        if "date" in lowered and ("amount" in lowered or "anount" in lowered):
            header_row_idx = row_idx
            for col_idx, cell in enumerate(cells, start=1):
                if cell.value is None:
                    continue
                header_map[str(cell.value).strip().lower()] = col_idx
            break
    return header_row_idx, header_map


def _next_empty_row(ws: Worksheet, start_row: int) -> int:
    row_idx = start_row
    while row_idx <= ws.max_row:
        row_cells = ws[row_idx]
        if not any(cell.value not in (None, "") for cell in row_cells):
            return row_idx
        row_idx += 1
    return ws.max_row + 1 if ws.max_row >= start_row else start_row


def _user_expenses_path(_: Optional[Path]) -> Path:
    """Always resolve to the profile's Expenses workbook, ensuring it exists."""
    _, expenses_path = ensure_user_files_exist()
    return expenses_path


def append_to_expenses_workbook(expenses_path: Path, data: Dict[str, object], lock_path: Optional[Path] = None) -> None:
    """Append a single expense row while preserving any existing formatting."""
    expenses_path = _user_expenses_path(expenses_path)
    lock = _lock_path(expenses_path, lock_path)
    with file_lock(lock):
        if expenses_path.exists():
            wb = load_workbook(expenses_path)
        else:
            wb = Workbook()
        if "Tracker" not in wb.sheetnames:
            raise ValueError("Sheet 'Tracker' not found in Expenses.xlsx")
        ws = wb["Tracker"]

        header_row_idx, header_map = _find_expense_headers(ws)
        if header_row_idx is None:
            # If the sheet looks empty, seed headers to keep structure stable.
            if ws.max_row < 1 or all(cell.value in (None, "") for cell in ws[1]):
                ws.append(["Date", "Type", "Description", "Payment Method", "Amount", "Notes"])
                header_row_idx = 1
                header_map = {"date": 1, "type": 2, "description": 3, "payment method": 4, "amount": 5, "notes": 6}
            else:
                header_row_idx = 1

        def _resolve_col(name_candidates: List[str], default_col: int) -> int:
            for cand in name_candidates:
                if cand in header_map:
                    return header_map[cand]
            return default_col

        rightmost = ws.max_column if ws.max_column else 1
        date_col = _resolve_col(["date"], 1)
        type_col = _resolve_col(["type"], max(rightmost + 1, 2))
        desc_col = _resolve_col(["description"], max(rightmost + 2, 3))
        pay_col = _resolve_col(["payment method", "payment_method"], max(rightmost + 3, 4))
        amount_col = _resolve_col(["amount", "anount"], max(rightmost + 4, 5))
        notes_col = _resolve_col(["notes", "note"], max(rightmost + 5, 6))

        # If we had to create new columns, label them on the header row to aid readability.
        for col_idx, name in [
            (date_col, "Date"),
            (type_col, "Type"),
            (desc_col, "Description"),
            (pay_col, "Payment Method"),
            (amount_col, "Amount"),
            (notes_col, "Notes"),
        ]:
            cell = ws.cell(row=header_row_idx, column=col_idx)
            if not cell.value:
                cell.value = name

        target_row = _next_empty_row(ws, start_row=header_row_idx + 1 if header_row_idx else 2)
        ws.cell(row=target_row, column=date_col, value=data.get("date"))
        ws.cell(row=target_row, column=type_col, value=data.get("type"))
        ws.cell(row=target_row, column=desc_col, value=data.get("description"))
        ws.cell(row=target_row, column=pay_col, value=data.get("payment_method"))
        ws.cell(row=target_row, column=amount_col, value=data.get("amount"))
        ws.cell(row=target_row, column=notes_col, value=data.get("notes"))

        _atomic_save(wb, expenses_path)


# -------------------- Reflections index support --------------------

def append_reflection_index(
    log_path: Path,
    lock_path: Optional[Path],
    date_for: str,
    docx_path: str,
    created_at: str,
    mood: Optional[int] = None,
) -> None:
    log_path = _as_path(log_path)
    headers = ["date", "docx_path", "created_at", "mood"]
    lock = _lock_path(log_path, lock_path)
    with file_lock(lock):
        wb = _load_or_create(log_path)
        ws = _ensure_sheet(wb, REFLECTIONS_INDEX_SHEET)
        header_map = _get_header_map(ws, headers)

        row_values = [""] * len(header_map)
        row_values[header_map["date"] - 1] = date_for
        row_values[header_map["docx_path"] - 1] = docx_path
        row_values[header_map["created_at"] - 1] = created_at
        if "mood" in header_map and mood is not None:
            row_values[header_map["mood"] - 1] = mood

        ws.append(row_values)
        _atomic_save(wb, log_path)


# -------------------- Expenses upsert support --------------------

def normalize_date(value: object) -> Optional[date]:
    """Convert common Excel/str date cells into date; returns None on failure."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial dates may appear as numbers; openpyxl usually converts but guard anyway.
        try:
            return datetime.fromordinal(date(1899, 12, 30).toordinal() + int(value)).date()
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    # Try a handful of common formats
    fmts = ["%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y", "%m/%d/%y"]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _find_header_row_generic(ws: Worksheet, max_scan: int = 20) -> tuple[Optional[int], Dict[str, int]]:
    """Scan first rows for a date header; returns (row_index, header_map)."""
    for row_idx in range(1, min(max_scan, ws.max_row) + 1):
        headers = [_norm_header(c.value) for c in ws[row_idx]]
        if any(h == "date" for h in headers):
            return row_idx, {h: idx + 1 for idx, h in enumerate(headers) if h}
    return None, {}


def upsert_daily_row(
    expenses_path: Path,
    entry_date: date,
    entry_type: str,
    amount: float,
    method: str,
    notes: str,
) -> None:
    """Fill or update the row for entry_date; accumulate amount and append notes."""
    expenses_path = _user_expenses_path(expenses_path)
    lock = _lock_path(expenses_path, None)
    with file_lock(lock):
        wb = load_workbook(expenses_path)
        if "Tracker" not in wb.sheetnames:
            raise ValueError("Sheet 'Tracker' not found in Expenses.xlsx")
        ws = wb["Tracker"]

        header_row_idx, header_map = _find_header_row_generic(ws)
        if not header_row_idx:
            raise ValueError("Could not find a header row with a Date column in Expenses.xlsx")

        def col(name: str, fallback: Optional[str] = None) -> Optional[int]:
            target = name
            if target in header_map:
                return header_map[target]
            if fallback and fallback in header_map:
                return header_map[fallback]
            return None

        date_col = col("date")
        type_col = col("type")
        notes_col = col("notes") or col("description")
        method_col = col("method") or col("payment_method")
        amount_col = col("amount") or col("anount")

        if not date_col or not amount_col:
            raise ValueError("Date or Amount column missing in Expenses.xlsx")

        today_row_idx: Optional[int] = None
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell_date = normalize_date(ws.cell(row=row_idx, column=date_col).value)
            if cell_date == entry_date:
                today_row_idx = row_idx
                break

        def _append_notes(existing: object, new_text: str) -> str:
            if not new_text:
                return str(existing) if existing is not None else ""
            if existing in (None, ""):
                return new_text
            return f"{existing} | {datetime.now().strftime('%H:%M')} - {new_text}"

        def _coerce_amount(val: object) -> float:
            try:
                if val is None or val == "":
                    return 0.0
                return float(val)
            except Exception:
                return 0.0

        if today_row_idx is None:
            # Fallback: append first empty row after data region.
            today_row_idx = ws.max_row + 1
            ws.cell(row=today_row_idx, column=date_col, value=entry_date)

        # Update row in-place.
        existing_type = ws.cell(row=today_row_idx, column=type_col).value if type_col else None
        existing_amount = _coerce_amount(ws.cell(row=today_row_idx, column=amount_col).value)
        existing_notes = ws.cell(row=today_row_idx, column=notes_col).value if notes_col else ""
        existing_method = ws.cell(row=today_row_idx, column=method_col).value if method_col else ""

        # Type handling: keep existing if set; only overwrite if empty.
        if type_col:
            if existing_type in (None, ""):
                ws.cell(row=today_row_idx, column=type_col, value=entry_type)
            elif str(existing_type).strip().lower() != str(entry_type).strip().lower():
                # Mismatched type -> append a note to avoid silent overwrite.
                notes = _append_notes(
                    existing_notes,
                    f"{entry_type} {amount} via {method}",
                )
                existing_notes = notes
                # Do not overwrite type; still accumulate amount separately below if matching.
            # else same type: proceed

        # Amount handling: accumulate if same type or type empty; else leave as-is.
        new_amount = existing_amount
        if existing_type in (None, "", entry_type):
            new_amount = existing_amount + float(amount)
        elif existing_amount == 0:
            new_amount = float(amount)
        ws.cell(row=today_row_idx, column=amount_col, value=new_amount)

        # Notes append
        if notes_col:
            ws.cell(row=today_row_idx, column=notes_col, value=_append_notes(existing_notes, notes))

        # Method set if empty
        if method_col:
            if existing_method in (None, ""):
                ws.cell(row=today_row_idx, column=method_col, value=method)
            elif str(existing_method).strip().lower() != str(method).strip().lower():
                # Preserve original; if notes exist, append method info
                if notes_col:
                    updated_notes = ws.cell(row=today_row_idx, column=notes_col).value
                    ws.cell(
                        row=today_row_idx,
                        column=notes_col,
                        value=_append_notes(updated_notes, f"Method: {method}"),
                    )

        _atomic_save(wb, expenses_path)
