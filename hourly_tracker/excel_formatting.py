from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
import traceback
from pathlib import Path
from typing import Dict, Iterable, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.bar_chart import BarChart
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .analytics import entries_to_blocks
from typing import Any
from .excel_store import (
    ENTRIES_COLUMNS,
    ENTRIES_SHEET,
    LOOKUP_COLUMNS,
    LOOKUP_SHEET,
    atomic_save_workbook,
    file_lock,
    load_or_create_workbook,
    read_categories,
    read_entries,
)

WEEKLY_CHART_SHEET = "Charts_Weekly"
DAILY_CHART_SHEET = "Charts_Daily"


def _style_headers(ws) -> None:
    header_fill = PatternFill(start_color="F4F6F8", end_color="F4F6F8", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")


def _apply_table(ws, name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    if ws.max_row < 2:
        return  # need at least header + one row for a valid table ref
    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    # Replace existing table safely without clobbering TableList with a plain list.
    for existing in list(ws._tables):
        if existing.displayName == name:
            ws._tables.remove(existing)
    try:
        ws.add_table(table)
    except AttributeError:
        # Some workbooks may have _tables as a plain list; append as a fallback.
        if isinstance(getattr(ws, "_tables", None), list):
            ws._tables.append(table)
        else:
            raise


def _freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_column >= 1 and ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"


def _category_validation(ws_entries, categories: List[str]) -> None:
    if not categories:
        return
    # Inline list validation to avoid fragile named ranges.
    cats_str = ",".join(sorted(set(categories)))
    dv = DataValidation(type="list", formula1=f'"{cats_str}"', allow_blank=True, showDropDown=True)
    ws_entries.add_data_validation(dv)

    try:
        cat_col_idx = ENTRIES_COLUMNS.index("category") + 1
    except ValueError:
        return
    start_cell = ws_entries.cell(row=2, column=cat_col_idx).coordinate
    end_cell = ws_entries.cell(row=max(2, ws_entries.max_row + 200), column=cat_col_idx).coordinate
    dv.add(f"{start_cell}:{end_cell}")


def _conditional_formatting(ws_entries) -> None:
    low_fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")

    try:
        focus_col = ENTRIES_COLUMNS.index("focus") + 1
        energy_col = ENTRIES_COLUMNS.index("energy") + 1
    except ValueError:
        return

    focus_range = f"{ws_entries.cell(row=2, column=focus_col).coordinate}:{ws_entries.cell(row=max(2, ws_entries.max_row + 200), column=focus_col).coordinate}"
    energy_range = f"{ws_entries.cell(row=2, column=energy_col).coordinate}:{ws_entries.cell(row=max(2, ws_entries.max_row + 200), column=energy_col).coordinate}"

    ws_entries.conditional_formatting.add(
        focus_range,
        CellIsRule(operator="lessThanOrEqual", formula=["2"], stopIfTrue=False, fill=low_fill),
    )
    ws_entries.conditional_formatting.add(
        energy_range,
        CellIsRule(operator="lessThanOrEqual", formula=["2"], stopIfTrue=False, fill=low_fill),
    )


def _weekly_category_matrix(cfg: Any, entries_override: Optional[List[dict]] = None) -> Tuple[List[str], List[str], Dict[Tuple[str, str], float]]:
    entries = entries_override if entries_override is not None else read_entries(cfg.log_path, lock_path=cfg.log_lock_path)
    blocks, _ = entries_to_blocks(entries, cfg)

    weeks: List[str] = []
    categories: List[str] = []
    matrix: Dict[Tuple[str, str], float] = defaultdict(float)

    for block in blocks:
        week_start_date = block.end.date() - timedelta(days=block.end.weekday())
        week_key = week_start_date.isoformat()
        if week_key not in weeks:
            weeks.append(week_key)
        if block.category not in categories:
            categories.append(block.category)
        matrix[(week_key, block.category)] += block.hours

    weeks.sort()
    categories.sort()
    return weeks, categories, matrix


def _build_weekly_chart_sheet(wb: Workbook, cfg: Any, entries_override: Optional[List[dict]] = None) -> None:
    if WEEKLY_CHART_SHEET in wb.sheetnames:
        del wb[WEEKLY_CHART_SHEET]
    ws = wb.create_sheet(WEEKLY_CHART_SHEET)

    weeks, categories, matrix = _weekly_category_matrix(cfg, entries_override=entries_override)
    if not weeks or not categories:
        ws.append(["No data yet"])
        return

    header = ["week_start"] + categories
    ws.append(header)
    for week in weeks:
        row = [week]
        for cat in categories:
            row.append(round(matrix.get((week, cat), 0.0), 2))
        ws.append(row)

    _style_headers(ws)
    _freeze_and_filter(ws)

    data_ref = Reference(ws, min_col=2, min_row=1, max_col=len(categories) + 1, max_row=len(weeks) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(weeks) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Weekly Time by Category"
    chart.y_axis.title = "Hours"
    chart.x_axis.title = "Week"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 24
    chart.height = 12
    chart.grouping = "stacked"

    ws.add_chart(chart, "H2")


def _daily_focus_series(cfg: Any, entries_override: Optional[List[dict]] = None) -> Tuple[List[str], List[float]]:
    entries = entries_override if entries_override is not None else read_entries(cfg.log_path, lock_path=cfg.log_lock_path)
    series: List[Tuple[str, float]] = []
    for row in entries:
        ts = row.get("timestamp")
        focus = row.get("focus")
        if not ts or focus is None:
            continue
        try:
            dt = datetime.fromisoformat(str(ts))
            series.append((dt.date().isoformat(), float(focus)))
        except Exception:
            continue

    if not series:
        return [], []

    by_day: Dict[str, List[float]] = defaultdict(list)
    for day, focus in series:
        by_day[day].append(focus)

    days = sorted(by_day.keys())
    avg_focus = [round(sum(vals) / max(1, len(vals)), 2) for vals in (by_day[d] for d in days)]
    return days, avg_focus


def _build_daily_chart_sheet(wb: Workbook, cfg: Any, entries_override: Optional[List[dict]] = None) -> None:
    if DAILY_CHART_SHEET in wb.sheetnames:
        del wb[DAILY_CHART_SHEET]
    ws = wb.create_sheet(DAILY_CHART_SHEET)

    days, avg_focus = _daily_focus_series(cfg, entries_override=entries_override)
    if not days:
        ws.append(["No data yet"])
        return

    ws.append(["date", "avg_focus"])
    for day, focus in zip(days, avg_focus):
        ws.append([day, focus])

    _style_headers(ws)
    _freeze_and_filter(ws)

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(days) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(days) + 1)

    chart = LineChart()
    chart.title = "Daily Average Focus"
    chart.y_axis.title = "Focus (1-5)"
    chart.x_axis.title = "Date"
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 24
    chart.height = 12

    ws.add_chart(chart, "D2")


def apply_excel_formatting(cfg: Any, entries_data: Optional[List[dict]] = None) -> None:
    cfg.resolve_paths()
    if cfg.log_path is not None and not isinstance(cfg.log_path, Path):
        cfg.log_path = Path(cfg.log_path)
    if getattr(cfg, "log_lock_path", None) is not None and not isinstance(cfg.log_lock_path, Path):
        cfg.log_lock_path = Path(cfg.log_lock_path)
    assert cfg.log_path is not None

    if entries_data is not None:
        categories = sorted({str(e.get("category") or "Other") for e in entries_data})
    else:
        categories = read_categories(cfg.log_path, lock_path=cfg.log_lock_path)

    lock_path = cfg.log_lock_path or cfg.log_path.with_suffix(cfg.log_path.suffix + ".lock")
    try:
        if entries_data is None:
            # Normal path with lock to avoid concurrent writes.
            with file_lock(lock_path):
                wb = load_or_create_workbook(cfg.log_path)
        else:
            # Avoid re-entrant lock when analytics already holds it.
            wb = load_or_create_workbook(cfg.log_path)

        ws_entries = wb[ENTRIES_SHEET] if ENTRIES_SHEET in wb.sheetnames else wb.create_sheet(ENTRIES_SHEET)
        ws_lookup = wb[LOOKUP_SHEET] if LOOKUP_SHEET in wb.sheetnames else wb.create_sheet(LOOKUP_SHEET)

        _style_headers(ws_entries)
        _style_headers(ws_lookup)
        _freeze_and_filter(ws_entries)
        _freeze_and_filter(ws_lookup)

        _category_validation(ws_entries, categories)
        _conditional_formatting(ws_entries)

        _apply_table(ws_entries, "EntriesTable")
        _apply_table(ws_lookup, "LookupTable")

        _build_weekly_chart_sheet(wb, cfg, entries_override=entries_data)
        _build_daily_chart_sheet(wb, cfg, entries_override=entries_data)

        atomic_save_workbook(wb, cfg.log_path)
    except (TimeoutError, PermissionError):
        try:
            log_path = cfg.state_dir / "app.log"
            log_path.parent.mkdir(parents=True, exist_ok=True)
            with log_path.open("a", encoding="utf-8") as fh:
                fh.write(f"[{datetime.now().isoformat(timespec='seconds')}] excel_formatting skipped: workbook locked\n")
        except Exception:
            pass
    except Exception:
        # Formatting should never block analytics/report generation; log and continue.
        try:
            log_path = cfg.state_dir / "app.log"
            log_path.parent.mkdir(parents=True, exist_ok=True)
            with log_path.open("a", encoding="utf-8") as fh:
                fh.write(f"[{datetime.now().isoformat(timespec='seconds')}] excel_formatting failed: {traceback.format_exc()}\n")
        except Exception:
            pass
